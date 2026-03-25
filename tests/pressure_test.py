#!/usr/bin/env python3
"""
RAG Pressure Test Harness

Programmatic test suite to measure indexing performance, query accuracy,
and edge case handling for the RAG Document Assistant.

Usage:
    python tests/pressure_test.py
    python tests/pressure_test.py --server-url http://localhost:8000
    python tests/pressure_test.py --skip-indexing  # if index already populated
"""

import argparse
import asyncio
import json
import os
import shutil
import time
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

import httpx
from openpyxl import load_workbook

# =============================================================================
# Configuration
# =============================================================================

DEFAULT_SERVER_URL = "http://localhost:8000"
TEST_DOCS_DIR = Path(__file__).parent / "ragPressureTestDocs"
TEMPLATE_PATH = Path(__file__).parent / "ragPressureTestReport.xlsx"
OUTPUT_DIR = Path(__file__).parent / "pressure_test_results"

# Batch sizes for indexing tests
BATCH_SIZES = [5, 10, 16]

# API timeouts (seconds)
INGEST_TIMEOUT = 300  # 5 min for large batches
SYNC_TIMEOUT = 300
QUERY_TIMEOUT = 60

# =============================================================================
# Data Models
# =============================================================================

@dataclass
class DocumentInfo:
    filename: str
    file_type: str
    size_kb: float
    path: str


@dataclass
class IndexingResult:
    batch_id: int
    batch_size: int
    documents: List[DocumentInfo]
    ingest_time_sec: float
    sync_time_sec: float
    total_time_sec: float
    chunks_created: int
    vectors_upserted: int
    errors: List[str]
    timestamp: str = field(default_factory=lambda: datetime.now().isoformat())


@dataclass
class QueryResult:
    query_id: int
    query_type: str
    query_text: str
    latency_ms: float
    answer: str
    citations: List[Dict[str, Any]]
    source_snippets: List[str]
    error: Optional[str] = None
    timestamp: str = field(default_factory=lambda: datetime.now().isoformat())


@dataclass
class EdgeCaseResult:
    test_id: int
    test_type: str
    query_text: str
    latency_ms: float
    answer: str
    pass_fail: str
    notes: str
    timestamp: str = field(default_factory=lambda: datetime.now().isoformat())


@dataclass
class PressureTestResults:
    test_run_id: str
    server_url: str
    start_time: str
    end_time: Optional[str]
    indexing_results: List[IndexingResult]
    query_results: List[QueryResult]
    edge_case_results: List[EdgeCaseResult]


# =============================================================================
# Logging
# =============================================================================

def log(message: str, level: str = "INFO"):
    """Print formatted log message."""
    timestamp = datetime.now().strftime("%H:%M:%S")
    prefix = {"INFO": "[INFO]", "OK": "[ OK ]", "WARN": "[WARN]", "ERR": "[ERR ]", ">>>": "[>>> ]"}
    print(f"{timestamp} {prefix.get(level, '[INFO]')} {message}")


# =============================================================================
# API Client Functions
# =============================================================================

async def check_health(client: httpx.AsyncClient, base_url: str) -> bool:
    """Check if server is running."""
    try:
        resp = await client.get(f"{base_url}/api/health", timeout=10)
        return resp.status_code == 200
    except Exception as e:
        log(f"Health check failed: {e}", "ERR")
        return False


async def clear_index(client: httpx.AsyncClient, base_url: str) -> Dict[str, Any]:
    """Clear Pinecone index."""
    try:
        resp = await client.delete(f"{base_url}/api/clear-index", timeout=60)
        return resp.json()
    except Exception as e:
        log(f"Clear index failed: {e}", "ERR")
        return {"error": str(e)}


async def ingest_documents(
    client: httpx.AsyncClient,
    base_url: str,
    docs_dir: str,
    output_path: str = "data/chunks.jsonl"
) -> Dict[str, Any]:
    """Ingest documents and return response with timing."""
    start = time.time()
    try:
        resp = await client.post(
            f"{base_url}/api/ingest",
            json={
                "docs_dir": docs_dir,
                "output_path": output_path,
                "provider": "bedrock"
            },
            timeout=INGEST_TIMEOUT
        )
        elapsed = time.time() - start
        result = resp.json()
        result["_elapsed_sec"] = elapsed
        return result
    except Exception as e:
        elapsed = time.time() - start
        log(f"Ingest failed: {e}", "ERR")
        return {"error": str(e), "_elapsed_sec": elapsed}


async def sync_pinecone(
    client: httpx.AsyncClient,
    base_url: str,
    chunks_path: str = "data/chunks.jsonl"
) -> Dict[str, Any]:
    """Sync vectors to Pinecone and return response with timing."""
    start = time.time()
    try:
        resp = await client.post(
            f"{base_url}/api/sync-pinecone",
            json={"chunks_path": chunks_path, "batch_size": 100},
            timeout=SYNC_TIMEOUT
        )
        elapsed = time.time() - start
        result = resp.json()
        result["_elapsed_sec"] = elapsed
        return result
    except Exception as e:
        elapsed = time.time() - start
        log(f"Sync failed: {e}", "ERR")
        return {"error": str(e), "_elapsed_sec": elapsed}


async def query_rag(
    client: httpx.AsyncClient,
    base_url: str,
    query: str,
    top_k: int = 3
) -> Dict[str, Any]:
    """Execute RAG query and return response with latency."""
    start = time.time()
    try:
        resp = await client.post(
            f"{base_url}/api/query",
            json={
                "query": query,
                "top_k": top_k,
                "use_hybrid": True,
                "use_reranking": True
            },
            timeout=QUERY_TIMEOUT
        )
        elapsed_ms = (time.time() - start) * 1000
        result = resp.json()
        result["_latency_ms"] = elapsed_ms
        return result
    except Exception as e:
        elapsed_ms = (time.time() - start) * 1000
        log(f"Query failed: {e}", "ERR")
        return {"error": str(e), "_latency_ms": elapsed_ms}


# =============================================================================
# Document Helpers
# =============================================================================

def get_document_list(docs_dir: Path) -> List[DocumentInfo]:
    """Get sorted list of documents with metadata."""
    docs = []
    for f in sorted(docs_dir.iterdir()):
        if f.is_file() and not f.name.startswith('.'):
            ext = f.suffix.lower().lstrip('.')
            size_kb = f.stat().st_size / 1024
            docs.append(DocumentInfo(
                filename=f.name,
                file_type=ext,
                size_kb=round(size_kb, 2),
                path=str(f)
            ))
    return docs


def prepare_batch_dir(docs: List[DocumentInfo], batch_num: int) -> Path:
    """Create temp directory with batch documents."""
    batch_dir = OUTPUT_DIR / f"batch_{batch_num}_docs"
    batch_dir.mkdir(parents=True, exist_ok=True)

    # Clear existing
    for f in batch_dir.iterdir():
        f.unlink()

    # Copy docs
    for doc in docs:
        shutil.copy(doc.path, batch_dir / doc.filename)

    return batch_dir


# =============================================================================
# Test Runners
# =============================================================================

async def run_indexing_tests(
    client: httpx.AsyncClient,
    base_url: str
) -> List[IndexingResult]:
    """Run indexing performance tests for all batch sizes."""
    results = []
    all_docs = get_document_list(TEST_DOCS_DIR)

    log(f"Found {len(all_docs)} documents in {TEST_DOCS_DIR}")

    for batch_num, batch_size in enumerate(BATCH_SIZES, 1):
        log(f"Batch {batch_num}/{len(BATCH_SIZES)}: Indexing {batch_size} documents", ">>>")

        # Get batch docs
        batch_docs = all_docs[:batch_size]

        # Clear index first
        log("Clearing index...")
        await clear_index(client, base_url)
        await asyncio.sleep(2)  # Wait for clear to propagate

        # Prepare batch directory
        batch_dir = prepare_batch_dir(batch_docs, batch_num)
        log(f"Prepared batch dir: {batch_dir}")

        # Ingest
        log("Running ingest...")
        ingest_result = await ingest_documents(client, base_url, str(batch_dir))
        ingest_time = ingest_result.get("_elapsed_sec", 0)

        if "error" in ingest_result and ingest_result.get("status") != "success":
            log(f"Ingest error: {ingest_result.get('error')}", "ERR")
        else:
            log(f"Ingest complete: {ingest_result.get('chunks', 0)} chunks in {ingest_time:.1f}s", "OK")

        # Sync to Pinecone
        log("Syncing to Pinecone...")
        sync_result = await sync_pinecone(client, base_url)
        sync_time = sync_result.get("_elapsed_sec", 0)

        if "error" in sync_result and sync_result.get("status") != "success":
            log(f"Sync error: {sync_result.get('error')}", "ERR")
        else:
            log(f"Sync complete: {sync_result.get('vectors_upserted', 0)} vectors in {sync_time:.1f}s", "OK")

        # Collect errors
        errors = []
        if ingest_result.get("errors"):
            errors.extend(ingest_result["errors"])
        if sync_result.get("errors"):
            errors.extend(sync_result["errors"])

        results.append(IndexingResult(
            batch_id=batch_num,
            batch_size=batch_size,
            documents=batch_docs,
            ingest_time_sec=round(ingest_time, 2),
            sync_time_sec=round(sync_time, 2),
            total_time_sec=round(ingest_time + sync_time, 2),
            chunks_created=ingest_result.get("chunks", 0),
            vectors_upserted=sync_result.get("vectors_upserted", 0),
            errors=errors
        ))

        log(f"Batch {batch_num} total: {ingest_time + sync_time:.1f}s", "OK")

    return results


async def run_query_tests(
    client: httpx.AsyncClient,
    base_url: str,
    queries: List[Dict[str, str]]
) -> List[QueryResult]:
    """Run query accuracy tests."""
    results = []

    log(f"Running {len(queries)} query tests", ">>>")

    for i, q in enumerate(queries, 1):
        query_text = q["query"]
        query_type = q.get("type", "unknown")

        log(f"Query {i}/{len(queries)}: {query_type}")

        response = await query_rag(client, base_url, query_text)

        latency = response.get("_latency_ms", 0)
        answer = response.get("answer", "")
        citations = response.get("citations", [])
        sources = response.get("sources", [])
        error = response.get("error")

        # Extract snippets from sources
        snippets = [s.get("snippet", "") for s in sources[:3]]

        if error:
            log(f"Query error: {error}", "WARN")
        else:
            log(f"Got answer ({latency:.0f}ms), {len(citations)} citations", "OK")

        results.append(QueryResult(
            query_id=i,
            query_type=query_type,
            query_text=query_text,
            latency_ms=round(latency, 1),
            answer=answer[:500] if answer else "",  # Truncate for storage
            citations=[{"id": c.get("id"), "score": c.get("score")} for c in citations],
            source_snippets=snippets,
            error=error
        ))

    return results


async def run_edge_case_tests(
    client: httpx.AsyncClient,
    base_url: str,
    edge_cases: List[Dict[str, str]]
) -> List[EdgeCaseResult]:
    """Run edge case tests."""
    results = []

    log(f"Running {len(edge_cases)} edge case tests", ">>>")

    for i, ec in enumerate(edge_cases, 1):
        query_text = ec["query"]
        test_type = ec.get("type", "unknown")
        expected = ec.get("expected", "")

        log(f"Edge case {i}/{len(edge_cases)}: {test_type}")

        # Special handling for post-disconnect test
        if test_type == "Post-disconnect":
            await clear_index(client, base_url)
            await asyncio.sleep(1)

        response = await query_rag(client, base_url, query_text)

        latency = response.get("_latency_ms", 0)
        answer = response.get("answer", "")
        error = response.get("error")

        # Determine pass/fail based on test type
        pass_fail = "Pass"
        notes = ""

        if test_type == "Out-of-scope":
            # Should indicate no relevant info or decline to answer
            if answer and ("don't have" in answer.lower() or
                          "no information" in answer.lower() or
                          "not in" in answer.lower() or
                          "cannot find" in answer.lower()):
                pass_fail = "Pass"
                notes = "Correctly indicated out-of-scope"
            elif answer and len(answer) > 50:
                pass_fail = "Fail"
                notes = "Hallucinated an answer"
            else:
                pass_fail = "Review"
                notes = "Manual review needed"

        elif test_type == "Multi-doc synthesis":
            # Should have multiple citations
            citations = response.get("citations", [])
            if len(citations) >= 2:
                pass_fail = "Pass"
                notes = f"Cited {len(citations)} sources"
            else:
                pass_fail = "Review"
                notes = f"Only {len(citations)} citation(s)"

        elif test_type == "Ambiguous":
            # Should handle gracefully (any reasonable response)
            if answer and not error:
                pass_fail = "Pass"
                notes = "Handled ambiguity"
            else:
                pass_fail = "Fail"
                notes = error or "No answer"

        elif test_type == "Very long query":
            # Should parse and respond
            if answer and len(answer) > 20:
                pass_fail = "Pass"
                notes = "Parsed long query successfully"
            else:
                pass_fail = "Fail"
                notes = "Failed to process"

        elif test_type == "Single-word":
            # Should infer intent
            if answer and not error:
                pass_fail = "Pass"
                notes = "Inferred intent"
            else:
                pass_fail = "Fail"
                notes = error or "No answer"

        elif test_type == "Post-disconnect":
            # Should fail gracefully
            if error or "no" in answer.lower() or not answer:
                pass_fail = "Pass"
                notes = "Handled missing index gracefully"
            else:
                pass_fail = "Review"
                notes = "Check if answer is valid"

        log(f"Result: {pass_fail} - {notes}", "OK" if pass_fail == "Pass" else "WARN")

        results.append(EdgeCaseResult(
            test_id=i,
            test_type=test_type,
            query_text=query_text[:200],
            latency_ms=round(latency, 1),
            answer=answer[:300] if answer else "",
            pass_fail=pass_fail,
            notes=notes
        ))

    return results


# =============================================================================
# Excel Output
# =============================================================================

def write_to_excel(
    results: PressureTestResults,
    template_path: Path,
    output_path: Path
):
    """Write results to Excel template."""
    log(f"Loading template: {template_path}")
    wb = load_workbook(template_path)

    # Tab 1: Indexing Performance
    if "Indexing Performance" in wb.sheetnames:
        ws = wb["Indexing Performance"]
        log("Writing to Indexing Performance tab")

        # Find data start row (skip headers)
        start_row = 2
        for ir in results.indexing_results:
            row = start_row + ir.batch_id - 1
            ws.cell(row=row, column=1, value=ir.batch_id)
            ws.cell(row=row, column=2, value=ir.batch_size)
            ws.cell(row=row, column=3, value=ir.ingest_time_sec)
            ws.cell(row=row, column=4, value=ir.sync_time_sec)
            ws.cell(row=row, column=5, value=ir.total_time_sec)
            ws.cell(row=row, column=6, value=ir.chunks_created)
            ws.cell(row=row, column=7, value=ir.vectors_upserted)
            # Avg time per doc
            avg_time = ir.total_time_sec / ir.batch_size if ir.batch_size > 0 else 0
            ws.cell(row=row, column=8, value=round(avg_time, 2))

    # Tab 2: Query Accuracy
    if "Query Accuracy" in wb.sheetnames:
        ws = wb["Query Accuracy"]
        log("Writing to Query Accuracy tab")

        start_row = 2
        for qr in results.query_results:
            row = start_row + qr.query_id - 1
            ws.cell(row=row, column=1, value=qr.query_id)
            ws.cell(row=row, column=2, value=qr.query_type)
            ws.cell(row=row, column=3, value=qr.query_text)
            ws.cell(row=row, column=4, value=qr.latency_ms)
            ws.cell(row=row, column=5, value=qr.answer[:500])
            ws.cell(row=row, column=6, value=json.dumps(qr.citations)[:500])
            ws.cell(row=row, column=7, value="; ".join(qr.source_snippets)[:500])
            # Leave columns 8-10 for manual scoring

    # Tab 3: Edge Cases
    if "Edge Cases" in wb.sheetnames:
        ws = wb["Edge Cases"]
        log("Writing to Edge Cases tab")

        start_row = 2
        for ec in results.edge_case_results:
            row = start_row + ec.test_id - 1
            ws.cell(row=row, column=1, value=ec.test_id)
            ws.cell(row=row, column=2, value=ec.test_type)
            ws.cell(row=row, column=3, value=ec.query_text)
            ws.cell(row=row, column=4, value=ec.latency_ms)
            ws.cell(row=row, column=5, value=ec.answer[:300])
            ws.cell(row=row, column=6, value=ec.pass_fail)
            ws.cell(row=row, column=7, value=ec.notes)

    # Save
    wb.save(output_path)
    log(f"Saved results to: {output_path}", "OK")


def save_json_results(results: PressureTestResults, output_dir: Path):
    """Save results as JSON backup."""
    output_dir.mkdir(parents=True, exist_ok=True)

    # Convert dataclasses to dicts
    data = {
        "test_run_id": results.test_run_id,
        "server_url": results.server_url,
        "start_time": results.start_time,
        "end_time": results.end_time,
        "indexing_results": [asdict(r) for r in results.indexing_results],
        "query_results": [asdict(r) for r in results.query_results],
        "edge_case_results": [asdict(r) for r in results.edge_case_results]
    }

    # Convert DocumentInfo objects in indexing results
    for ir in data["indexing_results"]:
        ir["documents"] = [asdict(d) if hasattr(d, '__dataclass_fields__') else d
                          for d in ir.get("documents", [])]

    json_path = output_dir / f"results_{results.test_run_id}.json"
    with open(json_path, 'w') as f:
        json.dump(data, f, indent=2, default=str)

    log(f"Saved JSON backup: {json_path}", "OK")


# =============================================================================
# Query Definitions
# =============================================================================

def get_test_queries() -> List[Dict[str, str]]:
    """Return the 10 test queries (matching Excel template)."""
    return [
        {"type": "Factual lookup", "query": "What are the two key variations of the parallelization workflow pattern according to Anthropic?"},
        {"type": "Cross-document", "query": "How do Anthropic, OpenAI, and Google each define the distinction between workflows and agents?"},
        {"type": "Synthesis", "query": "What workforce archetypes are predicted to emerge in the agent-orchestrated enterprise, and how do their roles differ?"},
        {"type": "Specific detail", "query": "In the Agentic AI Stack, what is the embedding dimensionality for sentence-transformers versus Nomic Embed?"},
        {"type": "Process/how-to", "query": "What are the five stages in the Claude decision framework for determining whether to build an agentic system?"},
        {"type": "No-answer (adversarial)", "query": "What is the maximum token limit for GPT-4 when used in agentic workflows?"},
        {"type": "Comparison", "query": "How does the sequential multi-agent pattern differ from the parallel multi-agent pattern in terms of orchestration?"},
        {"type": "Inference", "query": "Based on the AI Radar 2026 predictions, what challenges might organizations face when scaling AI agent deployments?"},
        {"type": "Edge case", "query": "What happens when a multi-agent loop pattern fails to meet its exit condition?"},
        {"type": "Vague/broad", "query": "Explain transformers"},
    ]


def get_edge_cases() -> List[Dict[str, str]]:
    """Return the 6 edge case tests (matching Excel template)."""
    return [
        {"type": "Out-of-scope", "query": "What is the best pizza topping for a Friday night dinner party?"},
        {"type": "Multi-doc synthesis", "query": "Compare the guardrail and safety mechanisms recommended by AWS, Microsoft, and NVIDIA for production AI agent deployments."},
        {"type": "Ambiguous", "query": "How do agents work?"},
        {"type": "Very long query", "query": "I am currently working on implementing an enterprise-grade AI agent system for my organization and I need to understand the complete architectural considerations including but not limited to the various design patterns available such as sequential processing, parallel execution, hierarchical coordination, and swarm-based approaches. Additionally, I want to know how these patterns compare across different vendor recommendations from major cloud providers and AI research organizations. Furthermore, I need guidance on decision frameworks that can help me determine when to use simple workflows versus fully autonomous agents, taking into account factors like task complexity, error tolerance, and the need for human oversight in critical decision points. Can you provide a comprehensive overview?"},
        {"type": "Single-word", "query": "Guardrails"},
        {"type": "Post-disconnect", "query": "What are the key agentic design patterns?"},
    ]


# =============================================================================
# Main
# =============================================================================

async def run_pressure_test(
    server_url: str,
    skip_indexing: bool = False,
    skip_queries: bool = False,
    skip_edge_cases: bool = False
) -> PressureTestResults:
    """Run the complete pressure test suite."""

    test_run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    start_time = datetime.now().isoformat()

    log(f"Starting pressure test run: {test_run_id}")
    log(f"Server URL: {server_url}")

    results = PressureTestResults(
        test_run_id=test_run_id,
        server_url=server_url,
        start_time=start_time,
        end_time=None,
        indexing_results=[],
        query_results=[],
        edge_case_results=[]
    )

    async with httpx.AsyncClient() as client:
        # Health check
        log("Checking server health...")
        if not await check_health(client, server_url):
            log("Server not responding! Is it running?", "ERR")
            return results
        log("Server is healthy", "OK")

        # Indexing tests
        if not skip_indexing:
            log("=" * 50)
            log("INDEXING PERFORMANCE TESTS")
            log("=" * 50)
            results.indexing_results = await run_indexing_tests(client, server_url)
        else:
            log("Skipping indexing tests (--skip-indexing)")

        # Ensure index is populated for query tests
        if not skip_indexing:
            # Re-index all docs for query/edge case tests
            log("Re-indexing all 16 docs for query tests...")
            all_docs = get_document_list(TEST_DOCS_DIR)
            batch_dir = prepare_batch_dir(all_docs, 99)
            await clear_index(client, server_url)
            await asyncio.sleep(2)
            await ingest_documents(client, server_url, str(batch_dir))
            await sync_pinecone(client, server_url)

        # Query accuracy tests
        if not skip_queries:
            log("=" * 50)
            log("QUERY ACCURACY TESTS")
            log("=" * 50)
            queries = get_test_queries()
            results.query_results = await run_query_tests(client, server_url, queries)
        else:
            log("Skipping query tests (--skip-queries)")

        # Edge case tests
        if not skip_edge_cases:
            log("=" * 50)
            log("EDGE CASE TESTS")
            log("=" * 50)
            edge_cases = get_edge_cases()
            results.edge_case_results = await run_edge_case_tests(client, server_url, edge_cases)
        else:
            log("Skipping edge case tests (--skip-edge-cases)")

    results.end_time = datetime.now().isoformat()
    return results


def main():
    parser = argparse.ArgumentParser(
        description="RAG Pressure Test Harness"
    )
    parser.add_argument(
        "--server-url",
        default=DEFAULT_SERVER_URL,
        help=f"API server URL (default: {DEFAULT_SERVER_URL})"
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=OUTPUT_DIR,
        help="Output directory for results"
    )
    parser.add_argument(
        "--skip-indexing",
        action="store_true",
        help="Skip indexing performance tests"
    )
    parser.add_argument(
        "--skip-queries",
        action="store_true",
        help="Skip query accuracy tests"
    )
    parser.add_argument(
        "--skip-edge-cases",
        action="store_true",
        help="Skip edge case tests"
    )

    args = parser.parse_args()

    # Ensure output dir exists
    args.output_dir.mkdir(parents=True, exist_ok=True)

    # Run tests
    results = asyncio.run(run_pressure_test(
        server_url=args.server_url,
        skip_indexing=args.skip_indexing,
        skip_queries=args.skip_queries,
        skip_edge_cases=args.skip_edge_cases
    ))

    # Save results
    log("=" * 50)
    log("SAVING RESULTS")
    log("=" * 50)

    # JSON backup
    save_json_results(results, args.output_dir)

    # Excel output
    output_xlsx = args.output_dir / f"ragPressureTestReport_{results.test_run_id}.xlsx"
    if TEMPLATE_PATH.exists():
        write_to_excel(results, TEMPLATE_PATH, output_xlsx)
    else:
        log(f"Template not found: {TEMPLATE_PATH}", "WARN")

    # Summary
    log("=" * 50)
    log("TEST COMPLETE")
    log("=" * 50)
    log(f"Indexing batches: {len(results.indexing_results)}")
    log(f"Query tests: {len(results.query_results)}")
    log(f"Edge cases: {len(results.edge_case_results)}")
    log(f"Results dir: {args.output_dir}")

    if results.query_results:
        avg_latency = sum(q.latency_ms for q in results.query_results) / len(results.query_results)
        log(f"Avg query latency: {avg_latency:.0f}ms")

    if results.edge_case_results:
        passed = sum(1 for e in results.edge_case_results if e.pass_fail == "Pass")
        log(f"Edge case pass rate: {passed}/{len(results.edge_case_results)}")


if __name__ == "__main__":
    main()
